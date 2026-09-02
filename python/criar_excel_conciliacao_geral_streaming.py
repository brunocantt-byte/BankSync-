from __future__ import annotations

from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
import csv
import os

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


BASE_DIR = Path(r"C:\ConciliaFinanceira\conciliados\ultimos_12_meses")
OUTPUT = BASE_DIR / "relatorio_conciliacao_ultimos_12_meses.xlsx"

SHEETS = [
    ("01_resumo.csv", "Painel", None),
    ("02_conciliados.csv", "Conciliados", None),
    ("03_valor_igual_validar.csv", "Valor Igual Validar", None),
    ("04_divergencias.csv", "Divergencias", None),
    ("05_banco_sem_sistema.csv", "Banco sem Sistema", None),
    ("06_sistema_sem_banco.csv", "Sistema sem Banco", 20000),
    ("07_duplicidades.csv", "Duplicidades", None),
    ("08_possiveis_divergencias_valor.csv", "Possiveis Divergencias", None),
]

CURRENCY_HEADERS = {"valor", "banco_valor", "sistema_valor", "diferenca_valor"}
DATE_HEADERS = {"data", "banco_data", "sistema_data"}
NUM_HEADERS = {"score", "quantidade"}

HEADER_FILL = PatternFill("solid", fgColor="18324A")
HEADER_FONT = Font(color="FFFFFF", bold=True)
HEADER_ALIGNMENT = Alignment(vertical="center", wrap_text=False)


def parse_cell(header: str, value: str):
    value = (value or "").replace("\r", " ").replace("\n", " ").strip()
    header = header.lower()
    if not value:
        return None
    if header in DATE_HEADERS and len(value) == 10 and value[4] == "-":
        try:
            return datetime.strptime(value, "%Y-%m-%d").date()
        except ValueError:
            return value
    if header in CURRENCY_HEADERS or header in NUM_HEADERS:
        try:
            return Decimal(value)
        except (InvalidOperation, ValueError):
            return value
    return value


def header_cell(ws, value):
    cell = WriteOnlyCell(ws, value=value)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = HEADER_ALIGNMENT
    return cell


def value_cell(ws, header, value):
    parsed = parse_cell(header, value)
    cell = WriteOnlyCell(ws, value=parsed)
    header = header.lower()
    cell.alignment = Alignment(wrap_text=False, vertical="center")
    if header in CURRENCY_HEADERS:
        cell.number_format = "#,##0.00"
    elif header in NUM_HEADERS:
        cell.number_format = "#,##0"
    elif header in DATE_HEADERS:
        cell.number_format = "yyyy-mm-dd"
    return cell


def set_widths(ws, headers):
    capped = {
        "banco_descricao": 55,
        "sistema_favorecido_descricao": 62,
        "descricao": 60,
        "observacao": 70,
        "favorecido": 38,
        "ids": 34,
        "chave": 34,
    }
    for index, header in enumerate(headers, start=1):
        letter = get_column_letter(index)
        ws.column_dimensions[letter].width = capped.get(header, min(max(len(header) + 4, 14), 28))


def append_csv_sheet(wb, filename, sheet_name, max_rows=None):
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    path = BASE_DIR / filename
    with path.open("r", encoding="utf-8-sig", newline="") as fp:
        reader = csv.reader(fp, delimiter=";")
        headers = next(reader, [])
        ws.append([header_cell(ws, header) for header in headers])
        set_widths(ws, headers)
        written_rows = 1
        for row_number, row in enumerate(reader, start=1):
            if max_rows is not None and row_number > max_rows:
                break
            ws.append([
                value_cell(ws, headers[index] if index < len(headers) else "", value)
                for index, value in enumerate(row)
            ])
            written_rows += 1
    if headers:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{written_rows}"
    return ws


def append_painel_notes(ws):
    rows = [
        ["Resumo Executivo", ""],
        ["Regra principal", "Valor, tipo, documento/CNPJ, data e relacao textual."],
        ["Valor igual", "Conciliado automaticamente e sinalizado para validacao."],
        ["Parcelas", "Somas exatas usadas quando havia evidencia por documento, fornecedor ou tributo Caixa."],
        ["Divergencias", "Possiveis diferencas de valor foram listadas quando documento/CNPJ raiz e periodo indicavam relacao."],
        ["Tributos Caixa", "Pagamentos a CAIXA ECONOMICA associados a impostos/encargos quando o Sistema indicava tributo."],
        ["Transferencias", "Mantidas na conciliacao bancaria; nao devem compor faturamento bruto."],
        ["Base Banco", os.getenv("BANKSYNC_PASTA_BANCO", "pasta oficial configurada localmente")],
        ["Gerado em", datetime.now().strftime("%Y-%m-%d %H:%M")],
        ["Observacao", "A aba Sistema sem Banco foi limitada a 20.000 linhas no Excel; o CSV completo fica na mesma pasta."],
    ]
    ws.append([])
    for index, row in enumerate(rows):
        if index == 0:
            ws.append([header_cell(ws, row[0]), header_cell(ws, row[1])])
        else:
            ws.append(row)


def main():
    wb = Workbook(write_only=True)
    for filename, sheet_name, max_rows in SHEETS:
        ws = append_csv_sheet(wb, filename, sheet_name, max_rows)
        if sheet_name == "Painel":
            append_painel_notes(ws)
    wb.save(OUTPUT)
    print(OUTPUT)


if __name__ == "__main__":
    main()
